"""
Com essa biblioteca será possível ler e escrever dados em células
específicas, formatar células, inserir gráficos,
criar fórmulas, adicionar imagens e outros elementos gráficos às suas
planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
Excel, como a criação de relatórios e análise de dados e/ou facilitando a
manipulação de grandes quantidades de informações.

Documentação: https://openpyxl.readthedocs.io/en/stable/
Instalação necessária: pip install  openpyxl
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

workbook = Workbook()
sheet_name = "Minha planilha"
workbook.create_sheet(sheet_name, 0)
worksheet: Worksheet = workbook[sheet_name]
workbook.remove(workbook["Sheet"])

worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    ["João", 14, 5.5],
    ["Maria", 13, 9.7],
    ["Luiz", 15, 8.8],
    ["Alberto", 16, 10],
]

# for i in range(1, 10):
#     for j in range(1, 4):
#         print(f"Linha: {i} | Coluna {j}")

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)
